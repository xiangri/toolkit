# -*- coding: utf-8 -*-

import os
import shutil
import fire
from fire.core import Fire
import openpyxl
import pandas as pd
from copy import deepcopy


def get_all_workbook_files_name(src_dir):
    workbook_files = []
    for root, dirs, files in os.walk(src_dir):
        for f in files:
            if f.endswith('.xlsx'):
                workbook_files.append(os.path.join(root, f))
                print("get ", f)
    return workbook_files


def check_row(row, not_null_columns):
    """
    全部not_null_columns列都不为空才返回True
    :param row:
    :param not_null_columns:
    :return:
    """
    for i in not_null_columns:
        if row[int(i)] is None:
            return False

    return True


def merge(src_dir, not_null_columns=None, result_file='all.xlsx'):
    workbook_files = get_all_workbook_files_name(src_dir)
    not_null_columns_dic = {}
    if not_null_columns is not None:
        for sheet_column in not_null_columns.split(','):
            sheet, column = sheet_column.split(':')
            if sheet and column:
                not_null_columns_dic[sheet] = int(column)

    wb_new = openpyxl.Workbook()

    for f in workbook_files:
        wb = openpyxl.load_workbook(f)
        for s in wb.worksheets:
            if s.title in wb_new.sheetnames:
                curr_sheet = wb_new[s.title]
            else:
                print("load sheet ", s.title)
                curr_sheet = wb_new.create_sheet(s.title)
            for row in s:
                # 如果表名在not_null_columns中，且这列的单元格值是空的就跳过这行
                if s.title in not_null_columns_dic and row[not_null_columns_dic[s.title]].value is None:
                    continue
                new_row = [cell.value for cell in row]
                curr_sheet.append(deepcopy(new_row))


    wb_new.save(result_file)
    print("merge done , result_file: ", result_file)


class Main(object):
    @staticmethod
    def merge2one(src_dir='.', not_null_columns=None, result_file='all.xlsx'):
        """
        not_null_columns='银行存款:1'
        """
        merge(src_dir, not_null_columns, result_file)
    
    @staticmethod
    def help():
        text = """
        1. 遍历src_dir，把所有同名的表合并到一个工作簿的一张里。
            --not_null_columns: 指定过滤项，不填即不过滤。比如 --not_null_columns='银行存款:1', 如果'银行存款'这张表的某一行第1(B)列的值为空就不要这一行
                                支持对每个表设置一列过滤列，--not_null_columns='银行存款:1,银行借款:3'
            --result_file: 结果保存到哪个文件，默认当前目录下的 all.xlsx

            useage : python merge_excel.py merge2one --src_dir='/Users/xiangri/github/toolkit/excel' --not_null_columns='银行存款:1,银行借款:3'
            easy   : python merge_excel.py merge2one '/Users/xiangri/github/toolkit/excel' '银行存款:1,银行借款:3'
        
        """
        print(text)


if __name__ == '__main__':
    # fire.Fire(Main)
    src_dir = r'c:\sss'
    not_null_columns = None
    # not_null_columns = '银行存款:1,银行借款:3'
    result_file = r'all.xlsx'
    Main.merge2one(src_dir, not_null_columns, result_file)
